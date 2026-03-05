"""
Export endpoints for CSV, Excel, and PDF reports
Enhanced with 40+ fields per machine
"""
import io
import csv
from typing import Optional, List
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, Response
import logging

from database import db_manager
from dependencies import get_optional_user, OptionalUser
from utils.adapters import adapt_document
from utils.helpers import calculate_machine_status
from utils.collections import (
    HEARTBEAT_LATEST, HARDWARE_LATEST, NETWORK_LATEST,
    SPECS_LATEST, ALERTS, MACHINE_ACTIONS, SESSIONS_LATEST
)

router = APIRouter(prefix="/api/v1", tags=["export"])
logger = logging.getLogger("university_monitoring")


async def get_comprehensive_machine_data():
    """Fetch comprehensive machine data with 40+ fields"""
    # Get all latest data
    heartbeat_cursor = db_manager.mongodb_db.heartbeat_monitor_latest.find({})
    heartbeat_data = {doc["machine_id"]: adapt_document(doc, "heartbeat") 
                      async for doc in heartbeat_cursor}
    
    hardware_cursor = db_manager.mongodb_db.hardware_monitor_latest.find({})
    hardware_data = {doc["machine_id"]: adapt_document(doc, "hardware") 
                     async for doc in hardware_cursor}
    
    network_cursor = db_manager.mongodb_db.network_monitor_latest.find({})
    network_data = {doc["machine_id"]: adapt_document(doc, "network") 
                    async for doc in network_cursor}
    
    specs_cursor = db_manager.mongodb_db.specs_monitor_latest.find({})
    specs_data = {doc["machine_id"]: adapt_document(doc, "specs") 
                  async for doc in specs_cursor}
    
    sessions_cursor = db_manager.mongodb_db.sessions_monitor_latest.find({})
    sessions_data = {doc["machine_id"]: adapt_document(doc, "sessions") 
                     async for doc in sessions_cursor}
    
    # Get notes count per machine
    from pymongo import ASCENDING
    notes_pipeline = [
        {"$group": {"_id": "$machine_id", "count": {"$sum": 1}}}
    ]
    notes_cursor = db_manager.mongodb_db.machine_notes.aggregate(notes_pipeline)
    notes_count = {doc["_id"]: doc["count"] async for doc in notes_cursor}
    
    # Get active alerts count
    alerts_pipeline = [
        {"$match": {"acknowledged": False}},
        {"$group": {"_id": "$machine_id", "count": {"$sum": 1}}}
    ]
    alerts_cursor = db_manager.mongodb_db.alerts.aggregate(alerts_pipeline)
    active_alerts = {doc["_id"]: doc["count"] async for doc in alerts_cursor}
    
    # Combine all data
    machines = []
    for machine_id, heartbeat in heartbeat_data.items():
        hw = hardware_data.get(machine_id, {})
        net = network_data.get(machine_id, {})
        spec = specs_data.get(machine_id, {})
        sess = sessions_data.get(machine_id, {})
        
        # Calculate uptime
        uptime_seconds = heartbeat.get("uptime_seconds", 0)
        uptime_days = round(uptime_seconds / 86400, 1) if uptime_seconds else 0
        
        # Calculate last seen
        last_seen = heartbeat.get("last_seen", "")
        if isinstance(last_seen, datetime):
            last_seen = last_seen.isoformat()
        
        # Session info
        current_user = sess.get("current_username", "")
        login_time = sess.get("login_time", "")
        if isinstance(login_time, datetime):
            login_time = login_time.isoformat()
        
        session_duration = sess.get("session_duration_seconds", 0)
        session_duration_hours = round(session_duration / 3600, 1) if session_duration else 0
        
        idle_time = sess.get("idle_time_seconds", 0)
        idle_time_min = round(idle_time / 60, 1) if idle_time else 0
        
        # Build comprehensive record (40+ fields)
        machine_record = {
            # === MACHINE INFO (5 fields) ===
            "machine_id": machine_id,
            "hostname": heartbeat.get("hostname", ""),
            "building": heartbeat.get("building", ""),
            "room": heartbeat.get("room", ""),
            "floor": heartbeat.get("floor", ""),
            
            # === STATUS (4 fields) ===
            "status": calculate_machine_status(heartbeat),
            "health_score": heartbeat.get("health_score", 0),
            "last_seen": last_seen,
            "uptime_days": uptime_days,
            
            # === CPU (5 fields) ===
            "cpu_model": spec.get("cpu_model", ""),
            "cpu_cores": spec.get("cpu_cores", 0),
            "cpu_freq_ghz": spec.get("cpu_frequency_ghz", 0),
            "cpu_usage_percent": hw.get("cpu_usage_percent", 0),
            "cpu_temp_c": hw.get("cpu_temperature_c", 0),
            
            # === MEMORY (4 fields) ===
            "ram_total_gb": hw.get("memory_total_gb", 0),
            "ram_usage_percent": hw.get("memory_usage_percent", 0),
            "ram_used_gb": hw.get("memory_used_gb", 0),
            "ram_free_gb": hw.get("memory_available_gb", 0),
            
            # === DISK (4 fields) ===
            "storage_type": spec.get("storage_type", ""),
            "storage_total_gb": hw.get("disk_total_gb", 0),
            "disk_usage_percent": hw.get("disk_usage_percent", 0),
            "disk_free_gb": round(hw.get("disk_total_gb", 0) - hw.get("disk_used_gb", 0), 1),
            
            # === NETWORK (5 fields) ===
            "primary_ip": net.get("primary_ip_address", ""),
            "mac_address": net.get("primary_mac_address", ""),
            "upload_mbps": net.get("upload_speed_mbps", 0),
            "download_mbps": net.get("download_speed_mbps", 0),
            "packet_loss_percent": net.get("packet_loss_percent", 0),
            
            # === OS (3 fields) ===
            "os_name": spec.get("os_name", ""),
            "os_version": spec.get("os_version", ""),
            "os_build": spec.get("os_build", ""),
            
            # === USER SESSION (4 fields) ===
            "current_user": current_user,
            "login_time": login_time,
            "session_duration_hours": session_duration_hours,
            "idle_time_min": idle_time_min,
            
            # === HARDWARE SPECS (4 fields) ===
            "gpu_model": spec.get("gpu_model", ""),
            "gpu_memory_gb": spec.get("gpu_memory_gb", 0),
            "motherboard": spec.get("motherboard_model", ""),
            "bios_version": spec.get("bios_version", ""),
            
            # === MANAGEMENT (3 fields) ===
            "notes_count": notes_count.get(machine_id, 0),
            "active_alerts": active_alerts.get(machine_id, 0),
            "last_maintenance": heartbeat.get("last_maintenance_date", ""),
        }
        
        machines.append(machine_record)
    
    return machines


@router.get("/export/machines")
async def export_machines(
    format: str = Query("csv", regex="^(csv|json|excel|xlsx)$"),
    building: Optional[str] = None,
    status: Optional[str] = None,
    user: OptionalUser = Depends(get_optional_user)
):
    """
    Export comprehensive machine list with 40+ fields
    
    Supports filtering by building and status
    Formats: csv, json, excel/xlsx
    """
    logger.info(f"Export machines requested: format={format}, building={building}, status={status}")
    
    # Get comprehensive data
    machines = await get_comprehensive_machine_data()
    
    # Apply filters
    if building:
        machines = [m for m in machines if m.get("building") == building]
    if status:
        machines = [m for m in machines if m.get("status") == status]
    
    # JSON format
    if format == "json":
        return {
            "machines": machines,
            "count": len(machines),
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "fields": len(machines[0].keys()) if machines else 0
        }
    
    # Excel format
    if format in ["excel", "xlsx"]:
        return await export_machines_excel(machines)
    
    # CSV format (default)
    output = io.StringIO()
    if machines:
        fieldnames = machines[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(machines)
    
    filename = f"machines-export-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def export_machines_excel(machines: List[dict]):
    """Export machines to Excel with multiple sheets and formatting"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed - falling back to CSV")
        # Fallback to CSV
        output = io.StringIO()
        if machines:
            writer = csv.DictWriter(output, fieldnames=machines[0].keys())
            writer.writeheader()
            writer.writerows(machines)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=machines.csv"}
        )
    
    wb = Workbook()
    
    # === SHEET 1: All Machines ===
    ws = wb.active
    ws.title = "All Machines"
    
    if machines:
        # Headers
        headers = list(machines[0].keys())
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for machine in machines:
            ws.append(list(machine.values()))
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # === SHEET 2: Summary ===
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Metric", "Value"])
    
    # Calculate summary stats
    total = len(machines)
    online = len([m for m in machines if m.get("status") == "online"])
    offline = len([m for m in machines if m.get("status") == "offline"])
    
    avg_cpu = sum(m.get("cpu_usage_percent", 0) for m in machines) / total if total else 0
    avg_memory = sum(m.get("ram_usage_percent", 0) for m in machines) / total if total else 0
    avg_disk = sum(m.get("disk_usage_percent", 0) for m in machines) / total if total else 0
    
    summary_ws.append(["Total Machines", total])
    summary_ws.append(["Online", online])
    summary_ws.append(["Offline", offline])
    summary_ws.append(["Avg CPU Usage %", round(avg_cpu, 1)])
    summary_ws.append(["Avg Memory Usage %", round(avg_memory, 1)])
    summary_ws.append(["Avg Disk Usage %", round(avg_disk, 1)])
    summary_ws.append(["Export Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    
    # Style summary headers
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 20
    
    # === SHEET 3: By Building ===
    if machines:
        building_ws = wb.create_sheet("By Building")
        building_ws.append(["Building", "Total", "Online", "Offline", "Avg CPU %", "Avg Memory %"])
        
        # Group by building
        buildings = {}
        for m in machines:
            bldg = m.get("building", "Unknown")
            if bldg not in buildings:
                buildings[bldg] = []
            buildings[bldg].append(m)
        
        for bldg, bldg_machines in sorted(buildings.items()):
            total_bldg = len(bldg_machines)
            online_bldg = len([m for m in bldg_machines if m.get("status") == "online"])
            offline_bldg = len([m for m in bldg_machines if m.get("status") == "offline"])
            avg_cpu_bldg = sum(m.get("cpu_usage_percent", 0) for m in bldg_machines) / total_bldg
            avg_mem_bldg = sum(m.get("ram_usage_percent", 0) for m in bldg_machines) / total_bldg
            
            building_ws.append([
                bldg,
                total_bldg,
                online_bldg,
                offline_bldg,
                round(avg_cpu_bldg, 1),
                round(avg_mem_bldg, 1)
            ])
        
        # Style headers
        for cell in building_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for col in range(1, 7):
            building_ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"machines-export-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/hardware-report")
async def export_hardware_report(
    format: str = Query("csv", regex="^(csv|json|excel|xlsx)$"),
    user: OptionalUser = Depends(get_optional_user)
):
    """Export detailed hardware specifications"""
    # Get comprehensive data
    machines = await get_comprehensive_machine_data()
    
    # Filter to hardware-relevant fields
    hardware_data = []
    for m in machines:
        hardware_data.append({
            "machine_id": m["machine_id"],
            "hostname": m["hostname"],
            "building": m["building"],
            "room": m["room"],
            "cpu_model": m["cpu_model"],
            "cpu_cores": m["cpu_cores"],
            "cpu_freq_ghz": m["cpu_freq_ghz"],
            "cpu_usage_percent": m["cpu_usage_percent"],
            "cpu_temp_c": m["cpu_temp_c"],
            "ram_total_gb": m["ram_total_gb"],
            "ram_usage_percent": m["ram_usage_percent"],
            "storage_type": m["storage_type"],
            "storage_total_gb": m["storage_total_gb"],
            "disk_usage_percent": m["disk_usage_percent"],
            "gpu_model": m["gpu_model"],
            "gpu_memory_gb": m["gpu_memory_gb"],
            "os_name": m["os_name"],
            "os_version": m["os_version"],
        })
    
    if format == "json":
        return {"hardware": hardware_data, "count": len(hardware_data)}
    
    if format in ["excel", "xlsx"]:
        return await export_excel_generic(hardware_data, "hardware-report")
    
    # CSV
    output = io.StringIO()
    if hardware_data:
        writer = csv.DictWriter(output, fieldnames=hardware_data[0].keys())
        writer.writeheader()
        writer.writerows(hardware_data)
    
    filename = f"hardware-report-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/alerts")
async def export_alerts(
    format: str = Query("csv", regex="^(csv|json|excel|xlsx)$"),
    acknowledged: Optional[bool] = None,
    severity: Optional[str] = None,
    user: OptionalUser = Depends(get_optional_user)
):
    """Export alerts with filtering"""
    query = {}
    if acknowledged is not None:
        query["acknowledged"] = acknowledged
    if severity:
        query["severity"] = severity
    
    from pymongo import DESCENDING
    cursor = db_manager.mongodb_db.alerts.find(query).sort("timestamp", DESCENDING)
    
    alerts = []
    async for doc in cursor:
        alerts.append({
            "machine_id": doc.get("machine_id", ""),
            "alert_type": doc.get("alert_type", ""),
            "severity": doc.get("severity", ""),
            "message": doc.get("message", ""),
            "timestamp": doc.get("timestamp").isoformat() if doc.get("timestamp") else "",
            "acknowledged": doc.get("acknowledged", False),
            "acknowledged_by": doc.get("acknowledged_by", ""),
            "acknowledged_at": doc.get("acknowledged_at").isoformat() if doc.get("acknowledged_at") else "",
            "created_at": doc.get("created_at").isoformat() if doc.get("created_at") else "",
        })
    
    if format == "json":
        return {"alerts": alerts, "count": len(alerts)}
    
    if format in ["excel", "xlsx"]:
        return await export_excel_generic(alerts, "alerts")
    
    # CSV
    output = io.StringIO()
    if alerts:
        writer = csv.DictWriter(output, fieldnames=alerts[0].keys())
        writer.writeheader()
        writer.writerows(alerts)
    
    filename = f"alerts-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf/system-report")
async def export_pdf_system_report(
    user: OptionalUser = Depends(get_optional_user)
):
    """Generate comprehensive PDF system report"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        logger.error("reportlab not installed")
        return {"error": "PDF generation not available. Install reportlab: pip install reportlab"}
    
    # Get data
    machines = await get_comprehensive_machine_data()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story.append(Paragraph("University Computer Monitoring System", title_style))
    story.append(Paragraph(f"System Report - {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", 
                          styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", styles['Heading2']))
    
    total = len(machines)
    online = len([m for m in machines if m.get("status") == "online"])
    offline = len([m for m in machines if m.get("status") == "offline"])
    avg_cpu = sum(m.get("cpu_usage_percent", 0) for m in machines) / total if total else 0
    avg_memory = sum(m.get("ram_usage_percent", 0) for m in machines) / total if total else 0
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Machines", str(total)],
        ["Online", str(online)],
        ["Offline", str(offline)],
        ["Uptime %", f"{(online/total*100):.1f}%" if total else "0%"],
        ["Avg CPU Usage", f"{avg_cpu:.1f}%"],
        ["Avg Memory Usage", f"{avg_memory:.1f}%"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Building breakdown
    story.append(Paragraph("Machines by Building", styles['Heading2']))
    
    buildings = {}
    for m in machines:
        bldg = m.get("building", "Unknown")
        if bldg not in buildings:
            buildings[bldg] = {"total": 0, "online": 0, "offline": 0}
        buildings[bldg]["total"] += 1
        if m.get("status") == "online":
            buildings[bldg]["online"] += 1
        else:
            buildings[bldg]["offline"] += 1
    
    building_data = [["Building", "Total", "Online", "Offline", "Uptime %"]]
    for bldg, stats in sorted(buildings.items()):
        uptime = (stats["online"] / stats["total"] * 100) if stats["total"] else 0
        building_data.append([
            bldg,
            str(stats["total"]),
            str(stats["online"]),
            str(stats["offline"]),
            f"{uptime:.1f}%"
        ])
    
    building_table = Table(building_data, colWidths=[1.8*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
    building_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    story.append(building_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Top issues
    story.append(Paragraph("Machines Requiring Attention", styles['Heading2']))
    
    issues = []
    for m in machines:
        if m.get("status") == "offline":
            issues.append([m["machine_id"], m["hostname"], "Offline", "Critical"])
        elif m.get("cpu_usage_percent", 0) > 90:
            issues.append([m["machine_id"], m["hostname"], "High CPU Usage", "Warning"])
        elif m.get("ram_usage_percent", 0) > 90:
            issues.append([m["machine_id"], m["hostname"], "High Memory Usage", "Warning"])
        elif m.get("disk_usage_percent", 0) > 90:
            issues.append([m["machine_id"], m["hostname"], "High Disk Usage", "Warning"])
    
    if issues:
        issue_data = [["Machine ID", "Hostname", "Issue", "Severity"]] + issues[:20]  # Top 20
        issue_table = Table(issue_data, colWidths=[1.5*inch, 2*inch, 2*inch, 1*inch])
        issue_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        story.append(issue_table)
    else:
        story.append(Paragraph("No critical issues found. All systems operating normally.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"system-report-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def export_excel_generic(data: List[dict], report_name: str):
    """Generic Excel export helper"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report_name}.csv"}
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = report_name.replace("-", " ").title()
    
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for row in data:
            ws.append(list(row.values()))
        
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{report_name}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
